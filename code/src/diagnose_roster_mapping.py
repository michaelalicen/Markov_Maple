#!/usr/bin/env python3
"""
diagnose_roster_mapping.py

Purpose
-------
Runs the current Scheduler.py solution, writes the workbook using OutputFormatter.py,
exports the raw CP-SAT selected assignments, and audits whether those raw assignments
actually appear in the final workbook.

This is designed to answer one question quickly:

    "Did Scheduler produce the assignment, and did OutputFormatter place it correctly?"

Default project layout expected:
    Markov_Maple/
      code/src/diagnose_roster_mapping.py
      code/src/Scheduler.py
      code/src/OutputFormatter.py
      data/vws_data.json
      data/workbook_clean.xlsx
      output/workbook_with_roster.xlsx

Usage from code/src:
    python diagnose_roster_mapping.py

Audit an already-produced workbook without solving again:
    python diagnose_roster_mapping.py --skip-solve \
      --raw-csv ../../output/mapping_audit/raw_solution_all.csv \
      --workbook ../../output/workbook_with_roster.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from datetime import date as DateType, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from ortools.sat.python import cp_model

import logger


FEASIBLE_STATUSES = {cp_model.OPTIMAL, cp_model.FEASIBLE}

VWS_SHEET = "VWS Roster"
CONTROL_SHEET = "Control Roster"
DISPATCH_SHEET = "Dispatch Roster"

RAW_FIELDS = ["roster", "date", "week", "base", "role", "volunteer_id"]


# ============================================================================
# Normalisation helpers
# ============================================================================

def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", _clean_text(value).upper())


def _norm_text(value: Any) -> str:
    return re.sub(r"\s+", " ", _clean_text(value).lower()).strip()


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _is_writeable_cell(cell) -> bool:
    return not isinstance(cell, MergedCell)


def _norm_base(value: Any) -> str:
    key = _norm_key(value)

    aliases = {
        "HDBHCSU": "HDBHCSU",
        "HDBHC": "HDBHCSU",
        "HDBSU": "HDBHCSU",
        "STBHC": "STBHC",
        "NWLHC1": "NWLHC1",
        "NWLHC2": "NWLHC2",
        "NWLSU": "NWLSU",
        "SPSHC": "SPSHC",
        "SPSBP": "SPSBP",
        "HELITACKSTB": "HELITACKSTB",
    }
    return aliases.get(key, key)


def _parse_date(value: Any) -> Optional[str]:
    """Return YYYY-MM-DD where possible."""
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, DateType):
        return value.strftime("%Y-%m-%d")

    text = _clean_text(value)
    if not text:
        return None

    # openpyxl reads formulas literally, e.g. '=B4+7'. Do not parse those here.
    if text.startswith("="):
        return None

    text = re.sub(r"\s+", " ", text).strip().rstrip(".")
    text = re.sub(r",\s*", ", ", text)

    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%A, %d %B %Y",
        "%A %d %B %Y",
        "%d %B %Y",
        "%d %b %Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass

    m = re.search(r"(\d{4}-\d{2}-\d{2})", text)
    if m:
        return m.group(1)

    return None


def _date_obj(value: Any) -> Optional[DateType]:
    parsed = _parse_date(value)
    if not parsed:
        return None
    try:
        return datetime.strptime(parsed, "%Y-%m-%d").date()
    except ValueError:
        return None


def _days_between(a: Any, b: Any) -> Optional[int]:
    da = _date_obj(a)
    db = _date_obj(b)
    if da is None or db is None:
        return None
    return abs((da - db).days)


def _role_category(role: Any) -> Optional[str]:
    r = _norm_key(role)

    if r in {"CL", "CREWLEADER"}:
        return "CL"
    if r in {"ACL", "ASSISTANTCL", "ASSISTANTCREWLEADER", "TRAINEEACL"}:
        return "ACL"
    if r in {"FF", "FIREFIGHTER", "RECRUITFF", "FFNR", "FF2YR", "NR"}:
        return "FF"
    if r in {"CREWDRIVER", "TRUCKDRIVER", "DRIVER"}:
        return "CREW_DRIVER"
    if r == "SKIDDRIVER":
        return "SKID_DRIVER"
    if r in {"LOGISTICS", "LOGISTICSSUPPORT", "LOGISTICSANDSUPPORT"}:
        return "LOGISTICS"
    if r == "PLANNING":
        return "PLANNING"
    if r in {"CONTROL", "CTRL"}:
        return "CONTROL"
    if r in {"MGR", "DISPATCHMANAGER", "MANAGER"}:
        return "DISPATCH_MANAGER"
    if r in {"NORM", "DISPATCHER", "NORMALDISPATCHER"}:
        return "DISPATCHER"
    if r in {"TRAINEE", "TRAINEEDISPATCHER"}:
        return "TRAINEE_DISPATCHER"

    return None


def _header_category(header: Any) -> Optional[str]:
    h = _norm_text(header)
    hk = _norm_key(header)

    if not h:
        return None

    if "assistant" in h or hk in {"ACL", "ASSISTANTCL", "ASSISTANTCREWLEADER"}:
        return "ACL"
    if "crew leader" in h or hk in {"CL", "CREWLEADER"}:
        return "CL"
    if "firefighter" in h or hk in {"FF", "FIREFIGHTER"}:
        return "FF"
    if "crew driver" in h or hk == "CREWDRIVER":
        return "CREW_DRIVER"
    if "skid driver" in h or hk == "SKIDDRIVER":
        return "SKID_DRIVER"
    if "truck driver" in h or hk == "TRUCKDRIVER":
        return "TRUCK_DRIVER"
    if "logistics" in h:
        return "LOGISTICS"
    if "planning" in h:
        return "PLANNING"
    if hk == "CONTROL" or h == "control":
        return "CONTROL"
    if "dispatch manager" in h or hk in {"MGR", "DISPATCHMANAGER"}:
        return "DISPATCH_MANAGER"
    if "trainee" in h and "dispatch" in h:
        return "TRAINEE_DISPATCHER"
    if h == "dispatcher" or hk in {"NORM", "DISPATCHER"}:
        return "DISPATCHER"

    return None


def _cell_contains_volunteer(cell_value: Any, volunteer_id: str) -> bool:
    return volunteer_id in _clean_text(cell_value)


# ============================================================================
# Project loading / solving
# ============================================================================

def _default_project_paths() -> Dict[str, Path]:
    src_dir = Path(__file__).resolve().parent
    root = src_dir.parent.parent

    return {
        "root": root,
        "data": root / "data" / "vws_data.json",
        "template": root / "data" / "workbook_clean.xlsx",
        "output": root / "output" / "workbook_with_roster.xlsx",
        "audit_dir": root / "output" / "mapping_audit",
    }


def load_data_file(path: Path):
    """Load either solver-ready JSON or cleaned workbook input."""
    if path.suffix.lower() == ".json":
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)

    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        # Load using data_cleaning.py in ../data_clean if available.
        src_dir = Path(__file__).resolve().parent
        data_clean_dir = src_dir.parent / "data_clean"
        if str(data_clean_dir) not in sys.path:
            sys.path.append(str(data_clean_dir))
        from data_cleaning import load_and_validate
        loaded = load_and_validate(path)
        return loaded[0] if isinstance(loaded, tuple) else loaded

    raise ValueError(f"Unsupported data file type: {path}")


def extract_raw_rows_from_solution(solution) -> Tuple[str, Dict[str, List[Dict[str, Any]]]]:
    solver, status, x, x_heli, x_ctrl, x_disp, model = solution
    status_name = solver.StatusName(status)

    if status not in FEASIBLE_STATUSES:
        raise RuntimeError(
            f"Cannot extract raw solution because CP-SAT returned {status_name}, "
            "not FEASIBLE/OPTIMAL."
        )

    rows: Dict[str, List[Dict[str, Any]]] = {
        "vws": [],
        "helitack": [],
        "control": [],
        "dispatch": [],
    }

    for (v, d, b, r), var in x.items():
        if solver.Value(var) == 1:
            rows["vws"].append({
                "roster": "vws",
                "date": d,
                "week": "",
                "base": b,
                "role": r,
                "volunteer_id": v,
            })

    for (v, d, r), var in x_heli.items():
        if solver.Value(var) == 1:
            rows["helitack"].append({
                "roster": "helitack",
                "date": d,
                "week": "",
                "base": "Helitack STB",
                "role": r,
                "volunteer_id": v,
            })

    for (v, w), var in x_ctrl.items():
        if solver.Value(var) == 1:
            rows["control"].append({
                "roster": "control",
                "date": "",
                "week": w,
                "base": "",
                "role": "Control",
                "volunteer_id": v,
            })

    for (v, w, r), var in x_disp.items():
        if solver.Value(var) == 1:
            rows["dispatch"].append({
                "roster": "dispatch",
                "date": "",
                "week": w,
                "base": "",
                "role": r,
                "volunteer_id": v,
            })

    for key in rows:
        rows[key].sort(key=lambda row: (
            str(row.get("date", "")),
            str(row.get("week", "")),
            str(row.get("base", "")),
            str(row.get("role", "")),
            str(row.get("volunteer_id", "")),
        ))

    return status_name, rows


def flatten_raw_rows(rows_by_roster: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for roster in ["vws", "helitack", "control", "dispatch"]:
        out.extend(rows_by_roster.get(roster, []))
    return out


def save_raw_solution(rows_by_roster: Dict[str, List[Dict[str, Any]]], out_dir: Path) -> Tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)

    raw_all = flatten_raw_rows(rows_by_roster)
    csv_path = out_dir / "raw_solution_all.csv"
    json_path = out_dir / "raw_solution_all.json"

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=RAW_FIELDS)
        writer.writeheader()
        for row in raw_all:
            writer.writerow({field: row.get(field, "") for field in RAW_FIELDS})

    with json_path.open("w", encoding="utf-8") as f:
        json.dump(rows_by_roster, f, indent=2)

    return csv_path, json_path


def load_raw_solution_csv(csv_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    rows: Dict[str, List[Dict[str, Any]]] = {
        "vws": [],
        "helitack": [],
        "control": [],
        "dispatch": [],
    }

    with csv_path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            roster = (row.get("roster") or "").strip().lower()
            if roster not in rows:
                continue
            rows[roster].append({field: row.get(field, "") for field in RAW_FIELDS})

    return rows


# ============================================================================
# Workbook scanning
# ============================================================================

def _find_table_header_row(ws, required_labels: Iterable[str], max_scan_rows: int = 60) -> int:
    required = {_norm_key(x) for x in required_labels}
    best_row = 1
    best_score = -1

    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        row_keys = {_norm_key(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)}
        score = sum(1 for label in required if label in row_keys)
        if score > best_score:
            best_score = score
            best_row = r

    return best_row


def _find_column_by_header(ws, header_row: int, possible_labels: Iterable[str]) -> Optional[int]:
    possible = {_norm_key(x) for x in possible_labels}
    for c in range(1, ws.max_column + 1):
        if _norm_key(ws.cell(row=header_row, column=c).value) in possible:
            return c
    return None


def _role_columns_from_header(ws, header_row: int) -> Dict[str, List[int]]:
    cols: Dict[str, List[int]] = {}
    for c in range(1, ws.max_column + 1):
        category = _header_category(ws.cell(row=header_row, column=c).value)
        if category:
            cols.setdefault(category, []).append(c)
    return cols


def _get_available_columns(role_columns: Dict[str, List[int]], category: Optional[str]) -> List[int]:
    if not category:
        return []

    cols = list(role_columns.get(category, []))

    if not cols and category == "TRUCK_DRIVER":
        cols = list(role_columns.get("CREW_DRIVER", []))
    if not cols and category == "CREW_DRIVER":
        cols = list(role_columns.get("TRUCK_DRIVER", []))
    if not cols and category == "LOGISTICS":
        cols = list(role_columns.get("PLANNING", []))
    if not cols and category == "PLANNING":
        cols = (
            list(role_columns.get("PLANNING", []))
            + list(role_columns.get("LOGISTICS", []))
            + list(role_columns.get("CREW_DRIVER", []))
        )

    return cols


def _build_vws_row_lookup(ws, header_row: int, date_col: int, base_col: int) -> Dict[Tuple[str, str], List[int]]:
    """Build (date, base) -> candidate row indices using carry-forward date/base."""
    lookup: Dict[Tuple[str, str], List[int]] = {}
    current_date: Optional[str] = None
    current_base: Optional[str] = None

    blank_run = 0

    for r in range(header_row + 1, ws.max_row + 1):
        parsed = _parse_date(ws.cell(row=r, column=date_col).value)
        if parsed:
            current_date = parsed
            current_base = None
            blank_run = 0

        base_cell = ws.cell(row=r, column=base_col).value
        base_key = _norm_base(base_cell) if not _is_blank(base_cell) else None
        if base_key:
            current_base = base_key
            blank_run = 0

        row_has_anything = any(not _is_blank(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1))

        if current_date and current_base and row_has_anything:
            lookup.setdefault((current_date, current_base), []).append(r)
            blank_run = 0
        else:
            blank_run += 1
            if blank_run >= 5:
                current_base = None

    # De-duplicate rows while preserving order.
    for key, row_list in list(lookup.items()):
        seen = set()
        lookup[key] = [rr for rr in row_list if not (rr in seen or seen.add(rr))]

    return lookup


def _candidate_vws_rows(row_lookup: Dict[Tuple[str, str], List[int]], date_value: str, base_value: str) -> List[int]:
    d = _parse_date(date_value) or str(date_value)
    b = _norm_base(base_value)

    exact = row_lookup.get((d, b), [])
    if exact:
        return exact

    # Defensive broad-base fallback.
    matches: List[int] = []
    for (dd, bb), rows in row_lookup.items():
        if dd != d:
            continue
        if bb.startswith(b) or b.startswith(bb):
            matches.extend(rows)
    return matches


def _find_exact_header_cell(ws, labels: Iterable[str], max_rows: int = 25) -> Optional[Tuple[int, int]]:
    targets = {_norm_key(x) for x in labels}
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, ws.max_column + 1):
            if _norm_key(ws.cell(row=r, column=c).value) in targets:
                return r, c
    return None


def _weekly_date_row_map(data_rows: List[int], ws, date_col: int, raw_weeks: Iterable[str]) -> Dict[str, int]:
    """Map week/date strings to workbook rows, inferring formula rows when needed."""
    mapping: Dict[str, int] = {}
    anchors: List[Tuple[int, DateType]] = []

    for idx, row_idx in enumerate(data_rows):
        parsed = _parse_date(ws.cell(row=row_idx, column=date_col).value)
        if parsed:
            try:
                anchors.append((idx, datetime.strptime(parsed, "%Y-%m-%d").date()))
            except ValueError:
                pass

    if anchors:
        anchor_idx, anchor_date = anchors[0]
        for idx, row_idx in enumerate(data_rows):
            inferred = anchor_date + timedelta(days=7 * (idx - anchor_idx))
            mapping[inferred.strftime("%Y-%m-%d")] = row_idx

        # Exact parsed dates override inferred rows.
        for idx, exact_date in anchors:
            mapping[exact_date.strftime("%Y-%m-%d")] = data_rows[idx]

    if not mapping:
        # Fallback: map sorted raw weeks to template rows in order.
        for week, row_idx in zip(sorted(set(raw_weeks)), data_rows):
            mapping[week] = row_idx

    return mapping


def _control_template_info(ws, raw_weeks: Iterable[str]) -> Tuple[Optional[int], Optional[int], Dict[str, int]]:
    found = _find_exact_header_cell(ws, ["Control"], max_rows=25)
    if not found:
        return None, None, {}

    header_row, control_col = found

    # Usually Week Starting is two columns left of Control.
    week_col = None
    for c in range(1, control_col):
        key = _norm_key(ws.cell(row=header_row, column=c).value)
        if key in {"WEEKSTARTING", "WEEKSTART", "START", "DATE", "WEEK"}:
            week_col = c
            break
    if week_col is None:
        week_col = max(1, control_col - 2)

    rows: List[int] = []
    for r in range(header_row + 1, ws.max_row + 1):
        label = _clean_text(ws.cell(row=r, column=1).value)
        if re.fullmatch(r"C\d+", label, flags=re.IGNORECASE):
            rows.append(r)

    if not rows:
        # Fallback: rows with week content/formulas under week_col
        for r in range(header_row + 1, ws.max_row + 1):
            if not _is_blank(ws.cell(row=r, column=week_col).value):
                rows.append(r)

    return control_col, week_col, _weekly_date_row_map(rows, ws, week_col, raw_weeks)


def _dispatch_template_info(ws, raw_weeks: Iterable[str]) -> Tuple[Dict[str, List[int]], Optional[int], Dict[str, int]]:
    role_header_row = None
    role_cols: Dict[str, List[int]] = {}
    best_score = 0

    for r in range(1, min(ws.max_row, 20) + 1):
        cols: Dict[str, List[int]] = {}
        for c in range(1, ws.max_column + 1):
            key = _norm_key(ws.cell(row=r, column=c).value)
            if key == "DISPATCHMANAGER":
                cols.setdefault("DISPATCH_MANAGER", []).append(c)
            elif key == "DISPATCHER":
                cols.setdefault("DISPATCHER", []).append(c)
            elif key == "TRAINEEDISPATCHER":
                cols.setdefault("TRAINEE_DISPATCHER", []).append(c)
        score = sum(len(v) for v in cols.values())
        if score > best_score:
            best_score = score
            role_header_row = r
            role_cols = cols

    if role_header_row is None or not role_cols:
        return {}, None, {}

    min_role_col = min(c for cols in role_cols.values() for c in cols)

    date_col = None
    for c in range(1, min_role_col):
        key = _norm_key(ws.cell(row=role_header_row, column=c).value)
        if key in {"WEEKSTARTING", "WEEKSTART", "START", "STARTDATE", "DATE", "WEEK"}:
            date_col = c
            break
    if date_col is None:
        date_col = 1

    data_rows: List[int] = []
    for r in range(role_header_row + 1, ws.max_row + 1):
        start_v = ws.cell(row=r, column=date_col).value
        if _parse_date(start_v) or (isinstance(start_v, str) and start_v.strip().startswith("=")) or not _is_blank(start_v):
            # exclude repeated header rows
            row_keys = {_norm_key(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)}
            if {"DISPATCHMANAGER", "DISPATCHER"} & row_keys:
                continue
            data_rows.append(r)

    return role_cols, date_col, _weekly_date_row_map(data_rows, ws, date_col, raw_weeks)


# ============================================================================
# Audit logic
# ============================================================================

def audit_vws(wb, raw_vws_rows: List[Dict[str, Any]], raw_helitack_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    ws = wb[VWS_SHEET]
    results: List[Dict[str, Any]] = []

    rows_to_check = list(raw_vws_rows) + [
        {
            **row,
            "roster": "helitack_in_vws",
            "base": "Helitack STB",
            "date": row.get("date", ""),
        }
        for row in raw_helitack_rows
    ]

    header_row = _find_table_header_row(ws, ["Date", "Base"])
    date_col = _find_column_by_header(ws, header_row, ["Date"])
    base_col = _find_column_by_header(ws, header_row, ["Base"])

    if date_col is None or base_col is None:
        for row in rows_to_check:
            results.append({
                **row,
                "sheet": VWS_SHEET,
                "status": "MISSING",
                "reason": "Could not detect Date/Base columns in VWS Roster",
                "candidate_rows": "",
                "candidate_cols": "",
                "found_cell": "",
            })
        return results

    role_cols = _role_columns_from_header(ws, header_row)
    row_lookup = _build_vws_row_lookup(ws, header_row, date_col, base_col)

    for row in rows_to_check:
        date_value = row.get("date", "")
        base_value = row.get("base", "")
        role_value = row.get("role", "")
        volunteer_id = row.get("volunteer_id", "")

        candidate_rows = _candidate_vws_rows(row_lookup, date_value, base_value)
        category = _role_category(role_value)
        candidate_cols = _get_available_columns(role_cols, category)

        found_cell = ""

        for rr in candidate_rows:
            for cc in candidate_cols:
                if _cell_contains_volunteer(ws.cell(row=rr, column=cc).value, volunteer_id):
                    found_cell = f"{ws.cell(row=rr, column=cc).coordinate}"
                    break
            if found_cell:
                break

        if found_cell:
            status = "FOUND"
            reason = ""
        elif not candidate_rows:
            status = "MISSING"
            reason = f"No candidate VWS row for date={date_value}, base={base_value}"
        elif not candidate_cols:
            status = "MISSING"
            reason = f"No candidate VWS role column for role={role_value} category={category}"
        else:
            status = "MISSING"
            reason = "Volunteer ID not found in candidate row/role cells"

        results.append({
            **row,
            "sheet": VWS_SHEET,
            "status": status,
            "reason": reason,
            "candidate_rows": ",".join(map(str, candidate_rows)),
            "candidate_cols": ",".join(map(str, candidate_cols)),
            "found_cell": found_cell,
        })

    return results


def audit_control(wb, raw_control_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    ws = wb[CONTROL_SHEET]
    results: List[Dict[str, Any]] = []

    raw_weeks = [(_parse_date(row.get("week")) or str(row.get("week", ""))) for row in raw_control_rows]
    control_col, week_col, date_to_row = _control_template_info(ws, raw_weeks)

    if control_col is None:
        for row in raw_control_rows:
            results.append({
                **row,
                "sheet": CONTROL_SHEET,
                "status": "MISSING",
                "reason": "Could not detect Control column",
                "candidate_rows": "",
                "candidate_cols": "",
                "found_cell": "",
            })
        return results

    for row in raw_control_rows:
        week_key = _parse_date(row.get("week")) or str(row.get("week", ""))
        volunteer_id = row.get("volunteer_id", "")
        row_idx = date_to_row.get(week_key)
        found_cell = ""

        if row_idx and _cell_contains_volunteer(ws.cell(row=row_idx, column=control_col).value, volunteer_id):
            found_cell = ws.cell(row=row_idx, column=control_col).coordinate

        results.append({
            **row,
            "sheet": CONTROL_SHEET,
            "status": "FOUND" if found_cell else "MISSING",
            "reason": "" if found_cell else f"Volunteer ID not found in Control cell for week={week_key}",
            "candidate_rows": str(row_idx or ""),
            "candidate_cols": str(control_col),
            "found_cell": found_cell,
        })

    return results


def audit_dispatch(wb, raw_dispatch_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    ws = wb[DISPATCH_SHEET]
    results: List[Dict[str, Any]] = []

    raw_weeks = [(_parse_date(row.get("week")) or str(row.get("week", ""))) for row in raw_dispatch_rows]
    role_cols, date_col, date_to_row = _dispatch_template_info(ws, raw_weeks)

    if not role_cols:
        for row in raw_dispatch_rows:
            results.append({
                **row,
                "sheet": DISPATCH_SHEET,
                "status": "MISSING",
                "reason": "Could not detect Dispatch role columns",
                "candidate_rows": "",
                "candidate_cols": "",
                "found_cell": "",
            })
        return results

    for row in raw_dispatch_rows:
        week_key = _parse_date(row.get("week")) or str(row.get("week", ""))
        role = row.get("role", "")
        volunteer_id = row.get("volunteer_id", "")
        category = _role_category(role)
        cols = _get_available_columns(role_cols, category)
        row_idx = date_to_row.get(week_key)
        found_cell = ""

        if row_idx:
            for cc in cols:
                if _cell_contains_volunteer(ws.cell(row=row_idx, column=cc).value, volunteer_id):
                    found_cell = ws.cell(row=row_idx, column=cc).coordinate
                    break

        results.append({
            **row,
            "sheet": DISPATCH_SHEET,
            "status": "FOUND" if found_cell else "MISSING",
            "reason": "" if found_cell else f"Volunteer ID not found in Dispatch role cell for week={week_key}, role={role}",
            "candidate_rows": str(row_idx or ""),
            "candidate_cols": ",".join(map(str, cols)),
            "found_cell": found_cell,
        })

    return results


def audit_workbook(raw_rows_by_roster: Dict[str, List[Dict[str, Any]]], workbook_path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(workbook_path, data_only=False)

    missing_sheets = [s for s in [VWS_SHEET, CONTROL_SHEET, DISPATCH_SHEET] if s not in wb.sheetnames]
    if missing_sheets:
        raise RuntimeError(f"Workbook is missing required roster sheets: {missing_sheets}")

    results: List[Dict[str, Any]] = []
    results.extend(audit_vws(wb, raw_rows_by_roster.get("vws", []), raw_rows_by_roster.get("helitack", [])))
    results.extend(audit_control(wb, raw_rows_by_roster.get("control", [])))
    results.extend(audit_dispatch(wb, raw_rows_by_roster.get("dispatch", [])))

    return results


def save_audit_results(results: List[Dict[str, Any]], out_dir: Path) -> Tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)

    audit_path = out_dir / "roster_mapping_audit.csv"
    missing_path = out_dir / "missing_assignments.csv"

    fields = RAW_FIELDS + ["sheet", "status", "reason", "candidate_rows", "candidate_cols", "found_cell"]

    with audit_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in results:
            writer.writerow({field: row.get(field, "") for field in fields})

    with missing_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in results:
            if row.get("status") != "FOUND":
                writer.writerow({field: row.get(field, "") for field in fields})

    return audit_path, missing_path


def print_summary(results: List[Dict[str, Any]]) -> None:
    print("\n" + "=" * 80)
    print("ROSTER MAPPING AUDIT SUMMARY")
    print("=" * 80)

    by_roster: Dict[str, Dict[str, int]] = {}

    for row in results:
        roster = str(row.get("roster", "unknown"))
        status = str(row.get("status", "unknown"))
        by_roster.setdefault(roster, {"FOUND": 0, "MISSING": 0, "OTHER": 0})
        if status in {"FOUND", "MISSING"}:
            by_roster[roster][status] += 1
        else:
            by_roster[roster]["OTHER"] += 1

    total_found = sum(v["FOUND"] for v in by_roster.values())
    total_missing = sum(v["MISSING"] for v in by_roster.values())

    for roster, counts in sorted(by_roster.items()):
        total = counts["FOUND"] + counts["MISSING"] + counts["OTHER"]
        print(
            f"{roster:16} total={total:4d} | "
            f"FOUND={counts['FOUND']:4d} | MISSING={counts['MISSING']:4d}"
        )

    print("-" * 80)
    print(f"Overall FOUND={total_found}, MISSING={total_missing}")

    missing = [row for row in results if row.get("status") != "FOUND"]
    if missing:
        print("\nFirst 20 missing assignments:")
        for row in missing[:20]:
            print(
                f"- {row.get('roster')} | date={row.get('date')} week={row.get('week')} "
                f"base={row.get('base')} role={row.get('role')} "
                f"volunteer={row.get('volunteer_id')} :: {row.get('reason')}"
            )
    else:
        print("\nAll raw solver assignments were found in the workbook.")


# ============================================================================
# CLI
# ============================================================================

def parse_args() -> argparse.Namespace:
    defaults = _default_project_paths()

    parser = argparse.ArgumentParser(
        description="Run/audit Scheduler raw assignments against the final populated workbook."
    )

    parser.add_argument("--data", type=Path, default=defaults["data"], help="Solver-ready JSON or cleaned workbook.")
    parser.add_argument("--template", type=Path, default=defaults["template"], help="Original clean workbook template.")
    parser.add_argument("--output", type=Path, default=defaults["output"], help="Populated output workbook.")
    parser.add_argument("--audit-dir", type=Path, default=defaults["audit_dir"], help="Folder to save raw/audit CSVs.")

    parser.add_argument(
        "--skip-solve",
        action="store_true",
        help="Do not solve/write. Audit an existing workbook against a previously saved raw CSV.",
    )
    parser.add_argument("--raw-csv", type=Path, default=None, help="Raw solution CSV for --skip-solve mode.")
    parser.add_argument("--workbook", type=Path, default=None, help="Workbook to audit for --skip-solve mode.")

    return parser.parse_args()


def main() -> int:
    args = parse_args()

    solver_log_path = args.audit_dir.parent / "solver_log.txt"
    args.audit_dir.parent.mkdir(parents=True, exist_ok=True)
    logger.setup_logger(solver_log_path)

    import os as _os
    try:
        _tty = open(_os.ctermid(), "w")
    except Exception:
        _tty = sys.stderr

    def tprint(*a, **kw):
        print(*a, **kw, file=_tty, flush=True)

    try:
        if args.skip_solve:
            if args.raw_csv is None:
                print("ERROR: --skip-solve requires --raw-csv")
                return 2
            workbook_path = args.workbook or args.output
            if not args.raw_csv.exists():
                print(f"ERROR: raw CSV not found: {args.raw_csv}")
                return 2
            if not workbook_path.exists():
                print(f"ERROR: workbook not found: {workbook_path}")
                return 2

            print(f"Auditing existing workbook: {workbook_path}")
            print(f"Against raw solution CSV:    {args.raw_csv}")

            raw_rows = load_raw_solution_csv(args.raw_csv)
            results = audit_workbook(raw_rows, workbook_path)
            audit_path, missing_path = save_audit_results(results, args.audit_dir)
            print_summary(results)
            print(f"\nSaved audit report:       {audit_path}")
            print(f"Saved missing assignments:{missing_path}")
            return 0 if all(row.get("status") == "FOUND" for row in results) else 1

        # Full run mode: solve -> export raw -> write workbook -> audit.
        print(f"Loading data:       {args.data}")
        print(f"Template workbook:  {args.template}")
        print(f"Output workbook:    {args.output}")
        print(f"Audit folder:       {args.audit_dir}")
        print(f"Solver log:         {solver_log_path}")

        from Scheduler import build_and_solve
        from OutputFormatter import write_roster

        data = load_data_file(args.data)

        tprint("Solver running (this takes ~100s)...")
        import time as _time
        _solve_start = _time.time()
        solution = build_and_solve(data)
        _solve_elapsed = _time.time() - _solve_start

        solver, status, *_ = solution
        status_name = solver.StatusName(status)

        tprint(f"Solution status : {status_name}  ({_solve_elapsed:.1f}s)")

        if status not in FEASIBLE_STATUSES:
            print("\nNo feasible/optimal solution returned.")
            print(f"CP-SAT status was: {status_name}")
            print("No raw solution/workbook audit can be performed.")
            tprint("No roster written. No audit performed.")
            return 2

        print(f"\nCP-SAT returned {status_name}. Extracting raw selected assignments...")
        status_name, raw_rows = extract_raw_rows_from_solution(solution)
        raw_csv, raw_json = save_raw_solution(raw_rows, args.audit_dir)

        print(f"Saved raw solution CSV:  {raw_csv}")
        print(f"Saved raw solution JSON: {raw_json}")

        tprint("Writing roster to workbook...")
        print("\nWriting workbook with OutputFormatter...")
        write_roster(solution, str(args.template), str(args.output))
        tprint(f"Roster written  : {args.output}")

        print("\nAuditing workbook against raw solution...")
        results = audit_workbook(raw_rows, args.output)
        audit_path, missing_path = save_audit_results(results, args.audit_dir)
        print_summary(results)

        print(f"\nSaved audit report:        {audit_path}")
        print(f"Saved missing assignments: {missing_path}")

        total_found   = sum(1 for r in results if r.get("status") == "FOUND")
        total_missing = sum(1 for r in results if r.get("status") == "MISSING")
        total_skipped = sum(1 for r in results if r.get("status") not in ("FOUND", "MISSING"))

        if total_missing > 0:
            print("\nConclusion: Scheduler produced these raw assignments, but some are missing in the workbook.")
            print("That points to OutputFormatter mapping/layout logic for the missing rows.")
            tprint(f"Audit result    : {total_found} FOUND, {total_missing} MISSING, {total_skipped} SKIPPED")
            tprint(f"Diagnostic log  : {solver_log_path}")
            tprint(f"Audit log       : {args.audit_dir.parent / 'diagnose_log.txt'}")
            return 1

        print("\nConclusion: every raw assignment was found in the workbook.")
        print("That means OutputFormatter mapping is consistent for this run.")
        tprint(f"Audit result    : {total_found} FOUND, 0 MISSING, {total_skipped} SKIPPED -- all assignments placed correctly")
        tprint(f"Diagnostic log  : {solver_log_path}")
        tprint(f"Audit log       : {args.audit_dir.parent / 'diagnose_log.txt'}")
        return 0

    finally:
        logger.close_logger()


if __name__ == "__main__":
    raise SystemExit(main())