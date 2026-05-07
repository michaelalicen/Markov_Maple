'''
Take the final CP-SAT solution and populate the existing SU/VWS workbook.

This version writes into the actual roster template/grid instead of replacing
it with a flat table. It also avoids appending fallback rows below the Control
and Dispatch templates.

It populates:
    - VWS Roster: by matching Date + Base row, then filling the correct role slot
    - Control Roster: by matching Week/Date row, then filling control role slots
    - Dispatch Roster: by matching Week/Date row, then filling dispatch role slots

It preserves the workbook layout as far as possible and only clears/writes the
assignment cells, not the date/base/heading structure.
'''

from __future__ import annotations

from datetime import date as DateType, datetime
from pathlib import Path
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from ortools.sat.python import cp_model

from logger import log_print


FEASIBLE_STATUSES = {cp_model.OPTIMAL, cp_model.FEASIBLE}

VWS_SHEET = "VWS Roster"
CONTROL_SHEET = "Control Roster"
DISPATCH_SHEET = "Dispatch Roster"
HELITACK_SHEET = "Helitack"

# Final output should only contain the roster result tabs.
# Helitack assignments are written into the VWS Roster as Helitack STB rows,
# so the standalone Helitack/source/cleaning sheets are removed from the saved copy.
OUTPUT_ROSTER_SHEETS = [VWS_SHEET, CONTROL_SHEET, DISPATCH_SHEET]

def _keep_only_roster_sheets(wb) -> None:
    """Remove all non-roster sheets from the output workbook.

    This is done after the roster tabs are populated, so the input template can
    still contain all source/cleaning sheets while the saved output is a clean
    deliverable containing only VWS Roster, Control Roster, and Dispatch Roster.
    """
    keep = set(OUTPUT_ROSTER_SHEETS)
    missing = [name for name in OUTPUT_ROSTER_SHEETS if name not in wb.sheetnames]
    if missing:
        raise ValueError(f"Cannot make roster-only workbook; missing required roster sheets: {missing}")

    for sheet_name in list(wb.sheetnames):
        if sheet_name not in keep:
            del wb[sheet_name]

    # Put the sheets in a predictable order.
    wb._sheets.sort(key=lambda ws: OUTPUT_ROSTER_SHEETS.index(ws.title))


# ---------------------------------------------------------------------------
# Solution extraction
# ---------------------------------------------------------------------------

def _sort_key(row: Dict[str, Any]) -> Tuple[str, ...]:
    return tuple(str(row.get(k, "")) for k in ["date", "week", "base", "role", "volunteer_id"])


def _selected_rows_from_solution(solution) -> Dict[str, List[Dict[str, Any]]]:
    '''
    Convert Scheduler.build_and_solve(...) output into plain row dictionaries.
    '''
    solver, status, x, x_heli, x_ctrl, x_disp, model = solution

    if status not in FEASIBLE_STATUSES:
        raise ValueError(
            "No feasible roster can be written because CP-SAT did not return "
            f"OPTIMAL or FEASIBLE. Status was: {solver.StatusName(status)}"
        )

    rows = {
        "vws": [],
        "helitack": [],
        "control": [],
        "dispatch": [],
    }

    for (v, d, b, r), var in x.items():
        if solver.Value(var) == 1:
            rows["vws"].append({
                "date": d,
                "base": b,
                "role": r,
                "volunteer_id": v,
            })

    for (v, d, r), var in x_heli.items():
        if solver.Value(var) == 1:
            rows["helitack"].append({
                "date": d,
                "role": r,
                "volunteer_id": v,
            })

    for (v, w), var in x_ctrl.items():
        if solver.Value(var) == 1:
            rows["control"].append({
                "week": w,
                "role": "Control",
                "volunteer_id": v,
            })

    for (v, w, r), var in x_disp.items():
        if solver.Value(var) == 1:
            rows["dispatch"].append({
                "week": w,
                "role": r,
                "volunteer_id": v,
            })

    for key in rows:
        rows[key] = sorted(rows[key], key=_sort_key)

    return rows


# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------

def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_text(value: Any) -> str:
    return re.sub(r"\s+", " ", _clean_text(value).lower())


def _norm_key(value: Any) -> str:
    '''Uppercase alphanumeric key. Good for matching base/header variants.'''
    return re.sub(r"[^A-Z0-9]", "", _clean_text(value).upper())


def _norm_base(value: Any) -> str:
    '''Normalise base names from JSON and workbook cells to the same key.'''
    key = _norm_key(value)

    aliases = {
        "HDBHCSU": "HDBHCSU",
        "HDBHC": "HDBHCSU",
        "HDBSU": "HDBHCSU",
        "STBHC": "STBHC",
        "NWLHC1": "NWLHC1",
        "NWLHC2": "NWLHC2",
        "NWLSU": "NWLSU",
        "SPSHC": "SPSHC",
        "SPSBP": "SPSBP",
        "HELITACKSTB": "HELITACKSTB",
    }

    return aliases.get(key, key)


def _parse_date(value: Any) -> Optional[str]:
    '''Return YYYY-MM-DD where possible.'''
    if value is None or value == "":
        return None

    if isinstance(value, (datetime, DateType)):
        return value.strftime("%Y-%m-%d")

    text = _clean_text(value)
    if not text:
        return None

    text = re.sub(r"\s+", " ", text).strip().rstrip(".")
    text = re.sub(r",\s*", ", ", text)

    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%A, %d %B %Y",
        "%A %d %B %Y",
        "%d %B %Y",
        "%d %b %Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass

    m = re.search(r"(\d{4}-\d{2}-\d{2})", text)
    if m:
        return m.group(1)

    return None


def _role_category(role: Any) -> Optional[str]:
    '''Map solver role names to the type of column that should be filled.'''
    r = _norm_key(role)

    if r in {"CL", "CREWLEADER"}:
        return "CL"
    if r in {"ACL", "ASSISTANTCL", "ASSISTANTCREWLEADER", "TRAINEEACL"}:
        return "ACL"
    if r in {"FF", "FIREFIGHTER", "RECRUITFF", "FFNR", "FF2YR", "NR"}:
        return "FF"
    if r in {"CREWDRIVER", "TRUCKDRIVER", "DRIVER"}:
        return "CREW_DRIVER"
    if r == "SKIDDRIVER":
        return "SKID_DRIVER"
    if r in {"LOGISTICS", "LOGISTICSSUPPORT", "LOGISTICSANDSUPPORT"}:
        return "LOGISTICS"
    if r == "PLANNING":
        return "PLANNING"
    if r in {"CONTROL", "CTRL"}:
        return "CONTROL"
    if r in {"MGR", "DISPATCHMANAGER", "MANAGER"}:
        return "DISPATCH_MANAGER"
    if r in {"NORM", "DISPATCHER", "NORMALDISPATCHER"}:
        return "DISPATCHER"
    if r in {"TRAINEE", "TRAINEEDISPATCHER"}:
        return "TRAINEE_DISPATCHER"

    return None


def _header_category(header: Any) -> Optional[str]:
    '''Map workbook header cells to role categories.'''
    h = _norm_text(header)
    hk = _norm_key(header)

    if not h:
        return None

    if "assistant" in h or hk in {"ACL", "ASSISTANTCL", "ASSISTANTCREWLEADER"}:
        return "ACL"
    if "crew leader" in h or hk in {"CL", "CREWLEADER"}:
        return "CL"
    if "firefighter" in h or hk in {"FF", "FIREFIGHTER"}:
        return "FF"
    if "crew driver" in h or hk == "CREWDRIVER":
        return "CREW_DRIVER"
    if "skid driver" in h or hk == "SKIDDRIVER":
        return "SKID_DRIVER"
    if "truck driver" in h or hk == "TRUCKDRIVER":
        return "TRUCK_DRIVER"
    if "logistics" in h:
        return "LOGISTICS"
    if "planning" in h:
        return "PLANNING"
    if "control" in h:
        return "CONTROL"
    if "dispatch manager" in h or "manager" in h or hk in {"MGR", "DISPATCHMANAGER"}:
        return "DISPATCH_MANAGER"
    if "trainee" in h and "dispatch" in h:
        return "TRAINEE_DISPATCHER"
    if "dispatcher" in h or hk in {"NORM", "DISPATCHER"}:
        return "DISPATCHER"

    return None


# ---------------------------------------------------------------------------
# Workbook scanning helpers
# ---------------------------------------------------------------------------

def _require_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"The workbook does not contain a sheet called '{sheet_name}'. "
            f"Available sheets are: {wb.sheetnames}"
        )
    return wb[sheet_name]


def _find_table_header_row(ws, required_labels: Iterable[str], max_scan_rows: int = 40) -> int:
    required = {_norm_key(x) for x in required_labels}
    best_row = 1
    best_score = -1

    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        row_keys = {_norm_key(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)}
        score = sum(1 for label in required if label in row_keys)
        if score > best_score:
            best_score = score
            best_row = r

    return best_row


def _find_column_by_header(ws, header_row: int, possible_labels: Iterable[str]) -> Optional[int]:
    possible = {_norm_key(x) for x in possible_labels}
    for c in range(1, ws.max_column + 1):
        if _norm_key(ws.cell(row=header_row, column=c).value) in possible:
            return c
    return None


def _is_writeable_cell(cell) -> bool:
    return not isinstance(cell, MergedCell)


def _range_intersects_area(cell_range, min_row: int, max_row: int, min_col: int, max_col: int) -> bool:
    return not (
        cell_range.max_row < min_row
        or cell_range.min_row > max_row
        or cell_range.max_col < min_col
        or cell_range.min_col > max_col
    )


def _unmerge_ranges_touching_columns(ws, start_row: int, columns: Iterable[int]) -> None:
    cols = sorted({int(c) for c in columns if c})
    if not cols:
        return

    for merged_range in list(ws.merged_cells.ranges):
        for col_idx in cols:
            if _range_intersects_area(
                merged_range,
                min_row=start_row,
                max_row=ws.max_row,
                min_col=col_idx,
                max_col=col_idx,
            ):
                ws.unmerge_cells(str(merged_range))
                break


def _clear_role_cells(ws, header_row: int, role_columns: Dict[str, List[int]]) -> None:
    all_cols = sorted({c for cols in role_columns.values() for c in cols})
    if not all_cols:
        return

    _unmerge_ranges_touching_columns(ws, header_row + 1, all_cols)

    for row_idx in range(header_row + 1, ws.max_row + 1):
        for col_idx in all_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            if _is_writeable_cell(cell):
                cell.value = None


def _get_available_columns(role_columns: Dict[str, List[int]], category: Optional[str]) -> List[int]:
    if not category:
        return []

    cols = list(role_columns.get(category, []))

    if not cols and category == "TRUCK_DRIVER":
        cols = list(role_columns.get("CREW_DRIVER", []))
    if not cols and category == "CREW_DRIVER":
        cols = list(role_columns.get("TRUCK_DRIVER", []))
    if not cols and category == "LOGISTICS":
        cols = list(role_columns.get("PLANNING", []))
    if not cols and category == "PLANNING":
        cols = (
            list(role_columns.get("PLANNING", []))
            + list(role_columns.get("LOGISTICS", []))
            + list(role_columns.get("CREW_DRIVER", []))
        )

    return cols


def _place_value_in_first_empty(ws, row_idx: int, columns: List[int], value: Any) -> bool:
    for col_idx in columns:
        cell = ws.cell(row=row_idx, column=col_idx)
        if not _is_writeable_cell(cell):
            continue
        if cell.value is None or str(cell.value).strip() == "":
            cell.value = value
            return True

    for col_idx in reversed(columns):
        cell = ws.cell(row=row_idx, column=col_idx)
        if not _is_writeable_cell(cell):
            continue
        existing = "" if cell.value is None else str(cell.value)
        cell.value = f"{existing}; {value}" if existing else value
        return True

    return False


def _write_to_empty_cell_only(ws, row_idx: int, columns: List[int], value: Any) -> bool:
    for col_idx in columns:
        cell = ws.cell(row=row_idx, column=col_idx)
        if not _is_writeable_cell(cell):
            continue
        if cell.value is None or str(cell.value).strip() == "":
            cell.value = value
            return True
    return False


def _append_value_to_last_writeable(ws, row_idx: int, columns: List[int], value: Any) -> bool:
    for col_idx in reversed(columns):
        cell = ws.cell(row=row_idx, column=col_idx)
        if not _is_writeable_cell(cell):
            continue
        existing = "" if cell.value is None else str(cell.value)
        cell.value = f"{existing}; {value}" if existing else value
        return True
    return False


def _place_value_across_candidate_rows(ws, row_indices: List[int], columns: List[int], value: Any) -> bool:
    for row_idx in row_indices:
        if _write_to_empty_cell_only(ws, row_idx, columns, value):
            return True

    for row_idx in reversed(row_indices):
        if _append_value_to_last_writeable(ws, row_idx, columns, value):
            return True

    return False


# ---------------------------------------------------------------------------
# IMPROVED VWS ROW LOOKUP - ONLY THIS PART WAS CHANGED
# ---------------------------------------------------------------------------

def _has_content_in_row(ws, row: int, start_col: int, look_ahead: int = 8) -> bool:
    """Check if row has any non-empty cells in the role area."""
    for c in range(start_col, min(start_col + look_ahead, ws.max_column + 1)):
        val = ws.cell(row=row, column=c).value
        if val is not None and str(val).strip() != "":
            return True
    return False


def _build_vws_row_lookup(ws, header_row: int, date_col: int, base_col: int) -> Dict[Tuple[str, str], List[int]]:
    '''Build (YYYY-MM-DD, base_key) -> list of row indices.

    Clean and robust version that properly handles continuation rows (e.g. SPS BP)
    without breaking the rest of the roster.
    '''
    lookup: Dict[Tuple[str, str], List[int]] = {}
    current_date: Optional[str] = None
    current_base: Optional[str] = None

    for r in range(header_row + 1, ws.max_row + 1):
        # Update date
        parsed = _parse_date(ws.cell(row=r, column=date_col).value)
        if parsed:
            current_date = parsed
            current_base = None

        # Update base
        base_value = ws.cell(row=r, column=base_col).value
        base_key = _norm_base(base_value) if _clean_text(base_value) else None
        if base_key:
            current_base = base_key

        if not current_date or not current_base:
            continue

        key = (current_date, current_base)

        include = False

        # 1. Normal row with content
        if _has_content_in_row(ws, r, base_col + 1):
            include = True
        # 2. Continuation row (immediately follows previous row of same block)
        elif key in lookup and lookup[key][-1] == r - 1:
            # Safety check - next row should not start new date/base
            next_r = r + 1
            if next_r <= ws.max_row:
                next_date = _parse_date(ws.cell(row=next_r, column=date_col).value)
                next_base_val = ws.cell(row=next_r, column=base_col).value
                next_base = _norm_base(next_base_val) if _clean_text(next_base_val) else None
                if not next_date and not next_base:
                    include = True

        if include:
            lookup.setdefault(key, []).append(r)

    # Deduplicate
    for key in lookup:
        lookup[key] = sorted(set(lookup[key]))

    return lookup


def _candidate_vws_rows(
    row_lookup: Dict[Tuple[str, str], List[int]],
    assignment_date: str,
    assignment_base: str,
) -> List[int]:
    date_key = _parse_date(assignment_date) or str(assignment_date)
    base_key = _norm_base(assignment_base)

    exact = row_lookup.get((date_key, base_key), [])
    if exact:
        return exact

    matches = []
    for (d, b), rows in row_lookup.items():
        if d == date_key and (b.startswith(base_key) or base_key.startswith(b)):
            matches.extend(rows)
    return sorted(set(matches))


def _write_vws_roster_grid(ws, rows: List[Dict[str, Any]]) -> List[str]:
    '''Populate the existing VWS Roster grid.'''
    warnings: List[str] = []

    header_row = _find_table_header_row(ws, ["Date", "Base"])
    date_col = _find_column_by_header(ws, header_row, ["Date"])
    base_col = _find_column_by_header(ws, header_row, ["Base"])

    if date_col is None or base_col is None:
        raise ValueError(
            f"Could not find Date/Base columns in '{ws.title}'. "
            "The formatter needs those columns to match CP-SAT assignments to template rows."
        )

    role_columns = _role_columns_from_header(ws, header_row)  # This function is defined in your full original file
    if not role_columns:
        raise ValueError(f"Could not detect role columns in '{ws.title}'.")

    _clear_role_cells(ws, header_row, role_columns)
    row_lookup = _build_vws_row_lookup(ws, header_row, date_col, base_col)

    for assignment in rows:
        a_date = str(assignment.get("date", ""))
        a_base = str(assignment.get("base", ""))
        a_role = assignment.get("role", "")
        volunteer_id = assignment.get("volunteer_id", "")

        candidate_rows = _candidate_vws_rows(row_lookup, a_date, a_base)
        if not candidate_rows:
            warnings.append(f"VWS assignment not written: no matching row for date={a_date}, base={a_base}, role={a_role}")
            continue

        category = _role_category(a_role)
        cols = _get_available_columns(role_columns, category)
        if not cols:
            warnings.append(f"VWS assignment not written: no matching role column for role={a_role}")
            continue

        written = _place_value_across_candidate_rows(ws, candidate_rows, cols, volunteer_id)
        if not written:
            warnings.append(f"Could not write to any cell for {a_date} {a_base} {a_role} {volunteer_id}")

    return warnings


# ===========================================================================
# The rest of the file (Control Roster, Dispatch Roster, Helitack, Public API, etc.)
# is 100% unchanged from your original pasted version.
# ===========================================================================

# [All your v7 overrides for Control and Dispatch, _write_control_roster_template,
# _write_dispatch_roster_template, write_roster, write_dummy_roster, etc. go here]

# Since your original paste was truncated, paste the rest of your original code 
# starting from where _write_vws_roster_grid ended in your file.

print("✅ OutputFormatter.py loaded with improved VWS row lookup")

# End of file